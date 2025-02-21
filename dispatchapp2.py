import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
import re
from tkinter import Tk, filedialog, messagebox

# Функція для перевірки часткової схожості
def is_partial_match(str1, str2, threshold=0.7):
    if not isinstance(str1, str) or not isinstance(str2, str):
        return False

    # Очищення рядків
    str1_clean = re.sub(r'[^\w]', '', str1).lower()
    str2_clean = re.sub(r'[^\w]', '', str2).lower()

    # Обчислення коефіцієнта схожості
    similarity = SequenceMatcher(None, str1_clean, str2_clean).ratio()
    return similarity >= threshold

# Функція для обробки файлів
def process_files():
    # Вибір файлів через GUI
    Tk().withdraw()  # Приховати основне вікно tkinter
    raport_file_path = filedialog.askopenfilename(title="Виберіть файл Raport_zwrot_postint", filetypes=[("Excel files", "*.xlsx")])
    zwroty_file_path = filedialog.askopenfilename(title="Виберіть файл Zwroty UA", filetypes=[("Excel files", "*.xlsx")])
    
    if not raport_file_path or not zwroty_file_path:
        messagebox.showwarning("Попередження", "Не вибрано обидва файли!")
        return

    # Завантаження даних
    raport_df = pd.read_excel(raport_file_path)
    zwroty_df = pd.read_excel(zwroty_file_path)

    # Очистка назв колонок
    raport_df.columns = raport_df.columns.str.strip()
    zwroty_df.columns = zwroty_df.columns.str.strip()

    # Пошук збігів
    matching_rows = []
    for raport_idx, raport_row in raport_df.iterrows():
        nr_przesylki_raport = raport_row["Nr_przesylki"]  # Колонка 1
        opis_towaru_raport = raport_row["Opis_towaru"]  # Колонка 8

        for _, zwroty_row in zwroty_df.iterrows():
            nr_przesylki_zwroty = zwroty_row["Nr_przesylki"]  # Колонка 2
            opis_towaru_zwroty = zwroty_row["Opis_towaru"]  # Колонка 5

            # Перевірка умов
            if (
                nr_przesylki_raport == nr_przesylki_zwroty
                and is_partial_match(opis_towaru_raport, opis_towaru_zwroty)
            ):
                matching_rows.append(raport_idx)
                break

    # Завантаження Excel для позначення кольором
    wb = load_workbook(raport_file_path)
    ws = wb.active

    # Стиль заповнення для виділення
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx in matching_rows:
        excel_row = row_idx + 2  # +2 через заголовок
        for cell in ws[excel_row]:
            cell.fill = highlight_fill

    # Збереження оновленого файлу
    output_file_path = filedialog.asksaveasfilename(title="Збережіть файл", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if output_file_path:
        wb.save(output_file_path)
        messagebox.showinfo("Успіх", f"Файл збережено за адресою:\n{output_file_path}")
    else:
        messagebox.showwarning("Попередження", "Файл не збережено!")

# Виконання функції
if __name__ == "__main__":
    process_files()
